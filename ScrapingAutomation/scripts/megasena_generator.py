from openpyxl import load_workbook, Workbook
from pprint import pprint
from random import sample
from time import sleep

file_path = 'C:\\Users\\Guilherme\\Documents\\testes\\home\\python\\ScrapingAutomation\\scripts\\results-converted.xlsx'
workbook = load_workbook(file_path)
ws = workbook.active

def iter_rows(workbook):
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            yield [cell.value for cell in row]

sequences = list(iter_rows(workbook))
#--------------------------------------
play_list = list()
play_num = int(input("Numero de jogos: "))

for i in range(play_num):
    sequence = sorted(sample(range(1, 61), 6))
    play_list.append(sequence[:])
    if sequence not in sequences:
        print(f'{i+1}: {sequence}')
    else:
        sequence = sorted(sample(range(1, 61), 6))
    sleep(0.5)

input('')










