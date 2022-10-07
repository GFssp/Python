from openpyxl import Workbook, load_workbook
from time import sleep

planilha_exe = load_workbook("C:/Users/jose marcos/Documents/Let's Code/Excel_python/controle_da_empresa (1).xlsx")

aba_v = planilha_exe['Vendas']

print('ACESSANDO PLANILHA EXCEL DE EXEMPLO: ')

for value in aba_v.iter_rows(values_only=True):
  print(value)
  sleep(0.5)